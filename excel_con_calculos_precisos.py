from openpyxl import Workbook
from datetime import datetime
import math

wb = Workbook()
wb.remove(wb.active)
año = datetime.now().year

# CÁLCULOS PRECISOS DE HIPOTECA
hipoteca_original = 3000000
hipoteca_tasa_anual = 3.25
hipoteca_plazo_años = 20
meses_transcurridos = 24

# Cálculo según la fórmula de la app
tipo_mensual = hipoteca_tasa_anual / 100 / 12
meses_totales = hipoteca_plazo_años * 12
meses_restantes = meses_totales - meses_transcurridos

# Capital pendiente (fórmula francesa)
factor_total = (1 + tipo_mensual) ** meses_totales
factor_transcurrido = (1 + tipo_mensual) ** meses_transcurridos
hipoteca_pendiente = hipoteca_original * (factor_total - factor_transcurrido) / (factor_total - 1)

# División CP/LP
factor_12_meses = (1 + tipo_mensual) ** (meses_transcurridos + 12)
capital_en_1_año = hipoteca_original * (factor_total - factor_12_meses) / (factor_total - 1)
hipoteca_cp = hipoteca_pendiente - capital_en_1_año
hipoteca_lp = capital_en_1_año

print(f"Hipoteca pendiente total: €{hipoteca_pendiente:,.2f}")
print(f"Porción CP: €{hipoteca_cp:,.2f}")
print(f"Porción LP: €{hipoteca_lp:,.2f}")

# CÁLCULOS DE PRÉSTAMO
prestamo_original = 500000
prestamo_tasa = 4.5
prestamo_plazo_años = 7
años_transcurridos = 2

# Similar cálculo para préstamo
tipo_anual = prestamo_tasa / 100
meses_prest = prestamo_plazo_años * 12
meses_trans_prest = años_transcurridos * 12
tipo_mensual_prest = tipo_anual / 12

factor_total_prest = (1 + tipo_mensual_prest) ** meses_prest
factor_trans_prest = (1 + tipo_mensual_prest) ** meses_trans_prest
prestamo_pendiente = prestamo_original * (factor_total_prest - factor_trans_prest) / (factor_total_prest - 1)

print(f"Préstamo pendiente: €{prestamo_pendiente:,.2f}")

# ACTIVOS OBJETIVO (basado en cálculos reales)
activo_corriente = 3000000
activo_fijo_bruto = 8000000
depreciacion = 1500000  # POSITIVO
intangibles_bruto = 600000
amortizacion = 100000  # POSITIVO
activo_no_corriente = (activo_fijo_bruto - depreciacion) + (intangibles_bruto - amortizacion)
total_activos = activo_corriente + activo_no_corriente

print(f"\nTotal Activos objetivo: €{total_activos:,.2f}")

# PASIVOS CALCULADOS
pasivo_corriente = 2000000  # Incluye líneas de crédito
pasivo_no_corriente = hipoteca_lp + prestamo_pendiente - (prestamo_pendiente * 0.2)  # Aproximación
total_pasivos_estimado = pasivo_corriente + pasivo_no_corriente + hipoteca_cp

print(f"Total Pasivos estimado: €{total_pasivos_estimado:,.2f}")

# PATRIMONIO PARA CUADRAR
patrimonio_necesario = total_activos - total_pasivos_estimado
print(f"Patrimonio necesario para cuadrar: €{patrimonio_necesario:,.2f}")

# CREAR EL EXCEL
# Info General
ws = wb.create_sheet("Info_General")
for row in [
    ["Campo", "Valor"],
    ["Nombre de la empresa", "Empresa Cálculos Precisos SA"],
    ["Sector", "Industrial"],
    ["País", "España"],
    ["Año de Fundación", 2015],
    ["¿Empresa familiar?", "No"],
    ["¿Cuentas auditadas?", "Sí"],
    ["Moneda", "EUR"]
]:
    ws.append(row)

# PYL
ws = wb.create_sheet("Datos_Historicos_PYL")
for row in [
    ["Concepto", str(año-2), str(año-1), str(año)],
    ["Ventas", 5000000, 5500000, 6000000],
    ["Costos Variables (%)", 45, 45, 45],
    ["Gastos de Personal", 1000000, 1050000, 1100000],
    ["Gastos Generales", 300000, 310000, 320000],
    ["Gastos de Marketing", 100000, 105000, 110000]
]:
    ws.append(row)

# Balance Activo
ws = wb.create_sheet("Balance_Activo")
for row in [
    ["Concepto", "Valor"],
    ["Tesorería", 500000],
    ["Inversiones financieras CP", 0],
    ["Clientes", 800000],
    ["Inventario", 1500000],
    ["Otros deudores", 100000],
    ["Administración Pública deudora", 50000],
    ["Gastos anticipados", 50000],
    ["Activos por impuesto diferido CP", 0],
    ["Activo fijo bruto", activo_fijo_bruto],
    ["Depreciación acumulada", depreciacion],  # POSITIVO
    ["Activos intangibles brutos", intangibles_bruto],
    ["Amortización acumulada intangibles", amortizacion],  # POSITIVO
    ["Inversiones financieras LP", 0],
    ["Créditos a largo plazo", 0],
    ["Fianzas y depósitos", 0],
    ["Activos por impuesto diferido LP", 0]
]:
    ws.append(row)

# Balance Pasivo
ws = wb.create_sheet("Balance_Pasivo")
for row in [
    ["Concepto", "Valor"],
    ["Proveedores", 800000],
    ["Acreedores por servicios", 100000],
    ["Anticipos de clientes", 50000],
    ["Remuneraciones pendientes", 100000],
    ["Administración Pública acreedora", 150000],
    ["Provisiones CP", 50000],
    ["Deuda financiera CP", 750000],  # Líneas de crédito
    ["Préstamos LP - Importe original", prestamo_original],
    ["Préstamos LP - Años transcurridos", años_transcurridos],
    ["Hipotecas - Importe original", hipoteca_original],
    ["Hipotecas - Meses transcurridos", meses_transcurridos],
    ["Leasing - Importe total", 100000],
    ["Leasing - Meses pendientes", 24],
    ["Otros préstamos LP", 0],
    ["Provisiones para riesgos LP", 50000],
    ["Pasivos por impuesto diferido", 25000]
]:
    ws.append(row)

# Patrimonio - ajustado para cuadrar
patrimonio_capital = 600000
patrimonio_reservas = 300000
patrimonio_resultados_acum = int(patrimonio_necesario - patrimonio_capital - patrimonio_reservas - 400000)
patrimonio_resultado = 400000

ws = wb.create_sheet("Balance_Patrimonio")
for row in [
    ["Concepto", "Valor"],
    ["Capital social", patrimonio_capital],
    ["Prima de emisión", 0],
    ["Reserva legal", 60000],
    ["Otras reservas", 240000],
    ["Resultados acumulados", patrimonio_resultados_acum],
    ["Resultado del ejercicio", patrimonio_resultado],
    ["Ajustes por cambios de valor", 0]
]:
    ws.append(row)

# Datos Laborales
ws = wb.create_sheet("Datos_Laborales")
for row in [
    ["Concepto", "Valor"],
    ["Número de empleados", 25],
    ["Coste medio por empleado (€/año)", 44000],
    ["Antigüedad media (años)", 3],
    ["Rotación anual (%)", 10]
]:
    ws.append(row)

# Líneas Financiación - deben sumar 750k
ws = wb.create_sheet("Lineas_Financiacion")
for row in [
    ["Tipo de Línea", "Banco", "Límite", "Dispuesto", "Tipo (%)", "Comisión (%)"],
    ["Póliza crédito", "Santander", 500000, 300000, 4.0, 0.5],
    ["Póliza crédito", "BBVA", 400000, 200000, 4.2, 0.5],
    ["Descuento comercial", "Sabadell", 300000, 150000, 3.8, 0.4],
    ["Póliza crédito stock", "CaixaBank", 200000, 100000, 3.5, 0.3]
]:
    ws.append(row)

# Parámetros
ws = wb.create_sheet("Proyecciones_Parametros")
for row in [
    ["Parámetro", "Valor"],
    ["Días de cobro (DSO)", 60],
    ["Días de pago (DPO)", 45],
    ["Días de inventario", 60],
    ["Crecimiento Año 1 (%)", 10],
    ["Crecimiento Año 2 (%)", 8],
    ["Crecimiento Año 3 (%)", 6],
    ["Crecimiento Año 4 (%)", 5],
    ["Crecimiento Año 5 (%)", 4]
]:
    ws.append(row)

wb.save('excel_calculos_precisos.xlsx')
print(f"\n✅ Excel creado con cálculos precisos")
print("Este Excel debería cuadrar cuando la app aplique las fórmulas de amortización francesa")
