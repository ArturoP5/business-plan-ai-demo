from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

wb = Workbook()
wb.remove(wb.active)

# Crear hoja Info_General con validaciones
ws = wb.create_sheet("Info_General")

# Encabezados
ws.append(["Campo", "Valor", "Opciones Válidas / Instrucciones"])

# Datos con instrucciones claras
datos = [
    ["Nombre de la empresa", "", "Texto libre"],
    ["Sector", "Tecnología", "Tecnología | Hostelería | Ecommerce | Consultoría | Retail | Servicios | Automoción | Industrial | Otro"],
    ["País", "España", "España | Francia | Portugal | Otros"],
    ["Año de Fundación", datetime.now().year - 5, "Año en formato YYYY"],
    ["¿Empresa familiar?", "No", "Sí | No (escribir exactamente)"],
    ["¿Cuentas auditadas?", "Sí", "Sí | No (escribir exactamente)"],
    ["Moneda", "EUR", "EUR | USD | GBP"]
]

for row_data in datos:
    ws.append(row_data)

# Agregar validaciones de datos (listas desplegables)
# Sector
dv_sector = DataValidation(type="list", 
                          formula1='"Tecnología,Hostelería,Ecommerce,Consultoría,Retail,Servicios,Automoción,Industrial,Otro"',
                          allow_blank=False)
dv_sector.add(ws['B3'])

# Sí/No para empresa familiar y cuentas auditadas
dv_si_no = DataValidation(type="list", formula1='"Sí,No"', allow_blank=False)
dv_si_no.add(ws['B6'])
dv_si_no.add(ws['B7'])

# Moneda
dv_moneda = DataValidation(type="list", formula1='"EUR,USD,GBP"', allow_blank=False)
dv_moneda.add(ws['B8'])

# Agregar las validaciones
ws.add_data_validation(dv_sector)
ws.add_data_validation(dv_si_no)
ws.add_data_validation(dv_moneda)

# Ajustar anchos
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 60

# Resaltar columna de instrucciones
for row in range(2, 9):
    ws[f'C{row}'].font = Font(italic=True, color="0066CC")

wb.save('test_validaciones.xlsx')
print("✅ Excel de prueba creado con validaciones")
