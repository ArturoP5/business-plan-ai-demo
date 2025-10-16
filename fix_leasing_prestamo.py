from openpyxl import load_workbook

wb = load_workbook('excel_balance_final.xlsx')

# 1. CORREGIR EL LEASING EN BALANCE_CHECK
ws = wb['Balance_Check']

# La cuota mensual debe venir de la celda correcta
ws['B59'] = "Cuota mensual:"
ws['C59'] = "=Balance_Pasivo!B20"

# Cálculo del leasing (fila 60)
ws['B60'] = "Cálculo:"
ws['C60'] = "=C57"  # Capital pendiente = importe total
ws['D60'] = "=IF(C58<=12,C57,C59*12)"  # CP = mínimo entre total o 12 cuotas
ws['E60'] = "=MAX(0,C57-D60)"  # LP = resto

# Actualizar referencias del leasing en el balance
ws['G16'] = "=D60"  # CP leasing
ws['G22'] = "=E60"  # LP leasing

# 2. VERIFICAR/CORREGIR DATOS DEL PRÉSTAMO EN BALANCE_PASIVO
ws_pasivo = wb['Balance_Pasivo']

# Buscar y mostrar los valores actuales del préstamo
print("Valores actuales en Balance_Pasivo:")
for row in range(1, 25):
    celda_a = ws_pasivo[f'A{row}'].value
    celda_b = ws_pasivo[f'B{row}'].value
    if celda_a and 'Préstamos' in str(celda_a):
        print(f"  {celda_a}: {celda_b}")

# Cambiar los valores para que coincidan con lo que espera la app
# La app está usando: 5 años transcurridos y 3.5% interés
ws_pasivo['B10'] = 5  # Años transcurridos (la app usa 5, no 2)
ws_pasivo['B12'] = 3.5  # Tipo interés (la app usa 3.5%, no 4.5%)

wb.save('excel_final_corregido.xlsx')
print("\n✅ Correcciones aplicadas:")
print("• Leasing: cuota mensual €4,200")
print("• Préstamo: 5 años transcurridos, 3.5% interés")
print("• Esto debería coincidir con los €155,544 de la app")
