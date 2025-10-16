"""
Plantilla Excel v2.0 - Versión definitiva para producción
Compatible 100% con la app, con Balance_Check profesional
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from io import BytesIO

def crear_plantilla_v2():
    """
    Crea la plantilla Excel v2.0 definitiva con Balance_Check profesional
    """
    wb = Workbook()
    wb.remove(wb.active)
    año_actual = datetime.now().year
    
    # Estilos
    titulo_style = Font(bold=True, size=14)
    header_style = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    subtotal_style = Font(bold=True)
    subtotal_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    
    # 1. INSTRUCCIONES
    ws = wb.create_sheet("INSTRUCCIONES")
    instrucciones = [
        ["PLANTILLA EXCEL v2.0 - BUSINESS PLAN IA"],
        [""],
        ["📋 PASOS PARA USAR ESTA PLANTILLA:"],
        [""],
        ["1. LEA ESTAS INSTRUCCIONES COMPLETAS"],
        ["2. Complete las hojas en orden:"],
        ["   • Info_General → Datos básicos de la empresa"],
        ["   • Datos_Historicos_PYL → Ventas y gastos últimos 3 años"],
        ["   • Balance_Activo → Todos los activos (depreciación en POSITIVO)"],
        ["   • Balance_Pasivo → Deudas y obligaciones"],
        ["   • Balance_Patrimonio → Capital y reservas"],
        ["   • Datos_Laborales → Información de empleados"],
        ["   • Lineas_Financiacion → Líneas de crédito activas"],
        ["   • Proyecciones_Parametros → Parámetros para proyecciones"],
        [""],
        ["3. VERIFIQUE EL BALANCE:"],
        ["   • Vaya a la hoja Balance_Check"],
        ["   • Debe mostrar '✅ BALANCE CUADRADO'"],
        ["   • Si no cuadra, revise los valores introducidos"],
        [""],
        ["⚠️ MUY IMPORTANTE:"],
        ["• Use 0 donde no haya datos (NO deje celdas vacías)"],
        ["• Depreciación y amortización en POSITIVO (la app los resta)"],
        ["• Guarde como .xlsx antes de cargar"],
        ["• NO modifique nombres de hojas"],
        [""],
        ["✅ VALIDACIÓN:"],
        ["• Balance_Check muestra el balance completo"],
        ["• Incluye cálculos de préstamos, hipotecas y leasing"],
        ["• Verifica automáticamente el cuadre"]
    ]
    for row_idx, instruccion in enumerate(instrucciones, 1):
        if instruccion:
            ws.cell(row=row_idx, column=1, value=instruccion[0])
            if row_idx == 1:
                ws.cell(row=row_idx, column=1).font = Font(bold=True, size=16, color="1E3A8A")
    ws.column_dimensions['A'].width = 80
    
    # 2. INFO GENERAL
    ws = wb.create_sheet("Info_General")
    for row in [
        ["Campo", "Valor", "Instrucciones"],
        ["Nombre de la empresa", "", "Razón social completa"],
        ["Sector", "Tecnología", "Tecnología/Industrial/Retail/Servicios/Otro"],
        ["País", "España", "España/Francia/Portugal/etc"],
        ["Año de Fundación", año_actual-5, "Año YYYY"],
        ["¿Empresa familiar?", "No", "Sí o No"],
        ["¿Cuentas auditadas?", "Sí", "Sí o No"],
        ["Moneda", "EUR", "EUR/USD/GBP"]
    ]:
        ws.append(row)
    # Agregar validaciones de datos (listas desplegables)
    dv_sector = DataValidation(type="list", 
                              formula1='"Tecnología,Hostelería,Ecommerce,Consultoría,Retail,Servicios,Automoción,Industrial,Otro"',
                              allow_blank=False)
    dv_sector.add(ws['B3'])
    
    dv_si_no = DataValidation(type="list", formula1='"Sí,No"', allow_blank=False)
    dv_si_no.add(ws['B6'])
    dv_si_no.add(ws['B7'])
    
    dv_moneda = DataValidation(type="list", formula1='"EUR,USD,GBP"', allow_blank=False)
    dv_moneda.add(ws['B8'])
    
    ws.add_data_validation(dv_sector)
    ws.add_data_validation(dv_si_no)
    ws.add_data_validation(dv_moneda)

    
    # 3. PYL HISTÓRICO
    ws = wb.create_sheet("Datos_Historicos_PYL")
    ws.append(["Concepto", str(año_actual-2), str(año_actual-1), str(año_actual)])
    ws.append(["Ventas", 0, 0, 0])
    ws.append(["Costos Variables (%)", 40, 40, 40])
    ws.append(["Gastos de Personal", 0, 0, 0])
    ws.append(["Gastos Generales", 0, 0, 0])
    ws.append(["Gastos de Marketing", 0, 0, 0])
    
    # 4. BALANCE ACTIVO
    ws = wb.create_sheet("Balance_Activo")
    ws.append(["Concepto", "Valor", "Notas"])
    ws.append(["Tesorería", 0, "Efectivo y equivalentes"])
    ws.append(["Inversiones financieras CP", 0, "Inversiones < 1 año"])
    ws.append(["Clientes", 0, "Cuentas por cobrar"])
    ws.append(["Inventario", 0, "Stock de productos"])
    ws.append(["Otros deudores", 0, "Otros deudores comerciales"])
    ws.append(["Administración Pública deudora", 0, "IVA, devoluciones"])
    ws.append(["Gastos anticipados", 0, "Seguros, alquileres prepagados"])
    ws.append(["Activos por impuesto diferido CP", 0, "Créditos fiscales CP"])
    ws.append(["Activo fijo bruto", 0, "Inmuebles, maquinaria, equipos"])
    ws.append(["Depreciación acumulada", 0, "⚠️ VALOR POSITIVO (la app resta)"])
    ws.append(["Activos intangibles brutos", 0, "Software, marcas, patentes"])
    ws.append(["Amortización acumulada intangibles", 0, "⚠️ VALOR POSITIVO (la app resta)"])
    ws.append(["Inversiones financieras LP", 0, "Inversiones > 1 año"])
    ws.append(["Créditos a largo plazo", 0, "Préstamos concedidos"])
    ws.append(["Fianzas y depósitos", 0, "Fianzas constituidas"])
    ws.append(["Activos por impuesto diferido LP", 0, "Créditos fiscales LP"])
    
    # 5. BALANCE PASIVO
    ws = wb.create_sheet("Balance_Pasivo")
    ws.append(["Concepto", "Valor", "Notas"])
    ws.append(["Proveedores", 0, "Facturas pendientes pago"])
    ws.append(["Acreedores por servicios", 0, "Otros acreedores"])
    ws.append(["Anticipos de clientes", 0, "Cobros anticipados"])
    ws.append(["Remuneraciones pendientes", 0, "Salarios pendientes"])
    ws.append(["Administración Pública acreedora", 0, "IRPF, SS pendientes"])
    ws.append(["Provisiones CP", 0, "Provisiones corto plazo"])
    ws.append(["Deuda financiera CP", 0, "⚠️ Debe coincidir con líneas dispuestas"])
    ws.append(["Préstamos LP - Importe original", 0, "Importe inicial préstamo"])
    ws.append(["Préstamos LP - Años transcurridos", 0, "Años ya pagados"])
    ws.append(["Préstamos LP - Plazo original (años)", 7, "Plazo total"])
    ws.append(["Préstamos LP - Tipo interés (%)", 4.5, "Tasa anual"])
    ws.append(["Préstamos LP - Comisión apertura (%)", 0.5, "% sobre importe"])
    ws.append(["Hipotecas - Importe original", 0, "Importe inicial hipoteca"])
    ws.append(["Hipotecas - Meses transcurridos", 0, "Meses ya pagados"])
    ws.append(["Hipotecas - Plazo total (años)", 20, "Plazo en años"])
    ws.append(["Hipotecas - Tipo interés (%)", 3.25, "Tasa anual"])
    ws.append(["Leasing - Importe total", 0, "Valor total del bien"])
    ws.append(["Leasing - Meses pendientes", 0, "Meses que faltan"])
    ws.append(["Leasing - Cuota mensual", 0, "Cuota mensual"])
    ws.append(["Otros préstamos LP", 0, "Otros préstamos bancarios"])
    ws.append(["Provisiones para riesgos LP", 0, "Contingencias LP"])
    ws.append(["Pasivos por impuesto diferido", 0, "Pasivos fiscales diferidos"])
    
    # 6. BALANCE PATRIMONIO
    ws = wb.create_sheet("Balance_Patrimonio")
    ws.append(["Concepto", "Valor", "Notas"])
    ws.append(["Capital social", 0, "Capital desembolsado"])
    ws.append(["Prima de emisión", 0, "Prima sobre nominal"])
    ws.append(["Reserva legal", 0, "Mínimo 10% capital social"])
    ws.append(["Otras reservas", 0, "Reservas voluntarias"])
    ws.append(["Resultados acumulados", 0, "Beneficios años anteriores"])
    ws.append(["Resultado del ejercicio", 0, "Resultado año actual"])
    ws.append(["Ajustes por cambios de valor", 0, "Ajustes valoración"])
    
    # 7. DATOS LABORALES
    ws = wb.create_sheet("Datos_Laborales")
    ws.append(["Concepto", "Valor", "Notas"])
    ws.append(["Número de empleados", 0, "Total empleados"])
    ws.append(["Coste medio por empleado (€/año)", 0, "Salario + SS empresa"])
    ws.append(["Antigüedad media (años)", 0, "Media años en empresa"])
    ws.append(["Rotación anual (%)", 0, "% rotación esperada"])
    
    # 8. LÍNEAS FINANCIACIÓN
    ws = wb.create_sheet("Lineas_Financiacion")
    ws.append(["Tipo de Línea", "Banco", "Límite", "Dispuesto", "Tipo (%)", "Comisión (%)"])
    # Líneas de ejemplo vacías
    for tipo in ["Póliza crédito", "Póliza crédito stock", "Descuento comercial"]:
        ws.append([tipo, "", 0, 0, 0, 0])
    
    # 9. PARÁMETROS
    ws = wb.create_sheet("Proyecciones_Parametros")
    ws.append(["Parámetro", "Valor", "Referencia"])
    ws.append(["Días de cobro (DSO)", 60, "Media del sector"])
    ws.append(["Días de pago (DPO)", 45, "Media del sector"])
    ws.append(["Días de inventario", 30, "Según tipo negocio"])
    ws.append(["Crecimiento Año 1 (%)", 10, "Sobre año actual"])
    ws.append(["Crecimiento Año 2 (%)", 8, "Sobre año 1"])
    ws.append(["Crecimiento Año 3 (%)", 6, "Sobre año 2"])
    ws.append(["Crecimiento Año 4 (%)", 5, "Sobre año 3"])
    ws.append(["Crecimiento Año 5 (%)", 4, "Sobre año 4"])
    ws.append(["CAPEX Año 1", 0, "Inversión en activos fijos"])
    ws.append(["CAPEX Año 2", 0, "Inversión en activos fijos"])
    ws.append(["CAPEX Año 3", 0, "Inversión en activos fijos"])
    ws.append(["CAPEX Año 4", 0, "Inversión en activos fijos"])
    ws.append(["CAPEX Año 5", 0, "Inversión en activos fijos"])    
    # 10. BALANCE_CHECK (la parte más importante)
    crear_balance_check_completo(wb, año_actual)
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def crear_balance_check_completo(wb, año_actual):
    """Crea la hoja Balance_Check con todas las fórmulas profesionales"""
    ws = wb.create_sheet("Balance_Check")
    
    # Título
    ws.merge_cells('B2:H2')
    ws['B2'] = f"BALANCE DE SITUACIÓN {año_actual} - VERIFICACIÓN"
    ws['B2'].font = Font(bold=True, size=14)
    ws['B2'].alignment = Alignment(horizontal='center')
    
    # El resto del código del Balance_Check aquí...
    # (Incluir todo el código de Balance_Check del excel_balance_final.xlsx)
    
    return ws

# Función para uso en la app
def descargar_plantilla_v2():
    """Función para ser llamada desde app.py"""
    return crear_plantilla_v2()

if __name__ == "__main__":
    excel = crear_plantilla_v2()
    with open("plantilla_v2_test.xlsx", "wb") as f:
        f.write(excel.getvalue())
    print("✅ Plantilla Excel v2.0 creada")

# Reemplazar la función crear_balance_check vacía con la completa
def crear_balance_check_completo(wb, año_actual):
    """Crea la hoja Balance_Check completa con todas las fórmulas"""
    ws = wb.create_sheet("Balance_Check")
    
    # Título
    ws.merge_cells('B2:H2')
    ws['B2'] = f"BALANCE DE SITUACIÓN {año_actual} - VERIFICACIÓN"
    ws['B2'].font = Font(bold=True, size=14)
    ws['B2'].alignment = Alignment(horizontal='center')
    
    # Encabezados principales
    ws['B4'] = "ACTIVO"
    ws['F4'] = "PASIVO Y PATRIMONIO NETO"
    for cell in ['B4', 'F4']:
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        ws[cell].alignment = Alignment(horizontal='center')
    
    # ACTIVO CORRIENTE
    ws['B6'] = "ACTIVO CORRIENTE"
    ws['B6'].font = Font(bold=True)
    ws['B7'] = "Tesorería"
    ws['C7'] = "=Balance_Activo!B2"
    ws['B8'] = "Inversiones financieras CP"
    ws['C8'] = "=Balance_Activo!B3"
    ws['B9'] = "Clientes"
    ws['C9'] = "=Balance_Activo!B4"
    ws['B10'] = "Inventario"
    ws['C10'] = "=Balance_Activo!B5"
    ws['B11'] = "Otros deudores"
    ws['C11'] = "=Balance_Activo!B6"
    ws['B12'] = "Admin. Pública deudora"
    ws['C12'] = "=Balance_Activo!B7"
    ws['B13'] = "Gastos anticipados"
    ws['C13'] = "=Balance_Activo!B8"
    ws['B14'] = "Activos impuesto diferido CP"
    ws['C14'] = "=Balance_Activo!B9"
    ws['B15'] = "Total Activo Corriente"
    ws['C15'] = "=SUM(C7:C14)"
    ws['B15'].font = Font(bold=True)
    
    # ACTIVO NO CORRIENTE
    ws['B17'] = "ACTIVO NO CORRIENTE"
    ws['B17'].font = Font(bold=True)
    ws['B18'] = "Activo fijo bruto"
    ws['C18'] = "=Balance_Activo!B10"
    ws['B19'] = "(-) Depreciación acumulada"
    ws['C19'] = "=-Balance_Activo!B11"
    ws['B20'] = "Activo fijo neto"
    ws['C20'] = "=C18+C19"
    ws['B21'] = "Intangibles brutos"
    ws['C21'] = "=Balance_Activo!B12"
    ws['B22'] = "(-) Amortización intangibles"
    ws['C22'] = "=-Balance_Activo!B13"
    ws['B23'] = "Intangibles netos"
    ws['C23'] = "=C21+C22"
    ws['B24'] = "Inversiones financieras LP"
    ws['C24'] = "=Balance_Activo!B14"
    ws['B25'] = "Créditos LP"
    ws['C25'] = "=Balance_Activo!B15"
    ws['B26'] = "Fianzas y depósitos"
    ws['C26'] = "=Balance_Activo!B16"
    ws['B27'] = "Activos impuesto diferido LP"
    ws['C27'] = "=Balance_Activo!B17"
    ws['B28'] = "Total Activo No Corriente"
    ws['C28'] = "=C20+C23+SUM(C24:C27)"
    ws['B28'].font = Font(bold=True)
    
    # TOTAL ACTIVO
    ws['B30'] = "TOTAL ACTIVO"
    ws['C30'] = "=C15+C28"
    ws['B30'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['B30'].fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    
    # PASIVO CORRIENTE
    ws['F6'] = "PASIVO CORRIENTE"
    ws['F6'].font = Font(bold=True)
    ws['F7'] = "Proveedores"
    ws['G7'] = "=Balance_Pasivo!B2"
    ws['F8'] = "Acreedores servicios"
    ws['G8'] = "=Balance_Pasivo!B3"
    ws['F9'] = "Anticipos clientes"
    ws['G9'] = "=Balance_Pasivo!B4"
    ws['F10'] = "Remuneraciones pendientes"
    ws['G10'] = "=Balance_Pasivo!B5"
    ws['F11'] = "Admin. Pública acreedora"
    ws['G11'] = "=Balance_Pasivo!B6"
    ws['F12'] = "Provisiones CP"
    ws['G12'] = "=Balance_Pasivo!B7"
    ws['F13'] = "Deuda financiera CP (líneas)"
    ws['G13'] = "=Balance_Pasivo!B8"
    ws['F14'] = "Préstamo CP (calculado)"
    ws['G14'] = "=D47"
    ws['F15'] = "Hipoteca CP (calculado)"
    ws['G15'] = "=D54"
    ws['F16'] = "Leasing CP (calculado)"
    ws['G16'] = "=D60"
    ws['F17'] = "Total Pasivo Corriente"
    ws['G17'] = "=SUM(G7:G16)"
    ws['F17'].font = Font(bold=True)
    
    # PASIVO NO CORRIENTE
    ws['F19'] = "PASIVO NO CORRIENTE"
    ws['F19'].font = Font(bold=True)
    ws['F20'] = "Préstamo LP (calculado)"
    ws['G20'] = "=E47"
    ws['F21'] = "Hipoteca LP (calculado)"
    ws['G21'] = "=E54"
    ws['F22'] = "Leasing LP (calculado)"
    ws['G22'] = "=E60"
    ws['F23'] = "Otros préstamos LP"
    ws['G23'] = "=Balance_Pasivo!B21"
    ws['F24'] = "Provisiones riesgos LP"
    ws['G24'] = "=Balance_Pasivo!B22"
    ws['F25'] = "Pasivos impuesto diferido"
    ws['G25'] = "=Balance_Pasivo!B23"
    ws['F26'] = "Total Pasivo No Corriente"
    ws['G26'] = "=SUM(G20:G25)"
    ws['F26'].font = Font(bold=True)
    
    # PATRIMONIO NETO
    ws['F28'] = "PATRIMONIO NETO"
    ws['F28'].font = Font(bold=True)
    ws['F29'] = "Capital social"
    ws['G29'] = "=Balance_Patrimonio!B2"
    ws['F30'] = "Prima de emisión"
    ws['G30'] = "=Balance_Patrimonio!B3"
    ws['F31'] = "Reserva legal"
    ws['G31'] = "=Balance_Patrimonio!B4"
    ws['F32'] = "Otras reservas"
    ws['G32'] = "=Balance_Patrimonio!B5"
    ws['F33'] = "Resultados acumulados"
    ws['G33'] = "=Balance_Patrimonio!B6"
    ws['F34'] = "Resultado ejercicio"
    ws['G34'] = "=Balance_Patrimonio!B7"
    ws['F35'] = "Ajustes por valor"
    ws['G35'] = "=Balance_Patrimonio!B8"
    ws['F36'] = "Total Patrimonio Neto"
    ws['G36'] = "=SUM(G29:G35)"
    ws['F36'].font = Font(bold=True)
    
    # TOTAL PASIVO + PN
    ws['F38'] = "TOTAL PASIVO + PATRIMONIO"
    ws['G38'] = "=G17+G26+G36"
    ws['F38'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['F38'].fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    
    # VERIFICACIÓN
    ws['B32'] = "VERIFICACIÓN:"
    ws['B32'].font = Font(bold=True, size=12, color="FF0000")
    ws['B33'] = "Diferencia (debe ser 0):"
    ws['C33'] = "=C30-G38"
    ws['B34'] = "Estado:"
    ws['C34'] = '=IF(ABS(C33)<1,"✅ BALANCE CUADRADO","❌ BALANCE DESCUADRADO")'
    
    # DETALLE CÁLCULOS DE DEUDA
    ws['B40'] = "DETALLE DE CÁLCULOS DE DEUDA (AMORTIZACIÓN FRANCESA)"
    ws['B40'].font = Font(bold=True, size=11)
    
    # PRÉSTAMO
    ws['B42'] = "PRÉSTAMO BANCARIO:"
    ws['C42'] = "Capital Pendiente"
    ws['D42'] = "Deuda CP"
    ws['E42'] = "Deuda LP"
    
    ws['B43'] = "Importe original:"
    ws['C43'] = "=Balance_Pasivo!B9"
    ws['B44'] = "Plazo (años):"
    ws['C44'] = "=Balance_Pasivo!B11"
    ws['B45'] = "Años transcurridos:"
    ws['C45'] = "=Balance_Pasivo!B10"
    ws['B46'] = "Tipo interés (%):"
    ws['C46'] = "=Balance_Pasivo!B12"
    
    ws['B47'] = "Cálculo:"
    ws['C47'] = "=IF(C43>0,C43*(((1+(C46/100)/12)^(C44*12))-((1+(C46/100)/12)^(C45*12)))/(((1+(C46/100)/12)^(C44*12))-1),0)"
    ws['D47'] = "=IF(C47>0,C47-C43*(((1+(C46/100)/12)^(C44*12))-((1+(C46/100)/12)^((C45+1)*12)))/(((1+(C46/100)/12)^(C44*12))-1),0)"
    ws['E47'] = "=MAX(0,C47-D47)"
    
    # HIPOTECA
    ws['B49'] = "HIPOTECA:"
    ws['C49'] = "Capital Pendiente"
    ws['D49'] = "Deuda CP"
    ws['E49'] = "Deuda LP"
    
    ws['B50'] = "Importe original:"
    ws['C50'] = "=Balance_Pasivo!B14"
    ws['B51'] = "Plazo (años):"
    ws['C51'] = "=Balance_Pasivo!B16"
    ws['B52'] = "Meses transcurridos:"
    ws['C52'] = "=Balance_Pasivo!B15"
    ws['B53'] = "Tipo interés (%):"
    ws['C53'] = "=Balance_Pasivo!B17"
    
    ws['B54'] = "Cálculo:"
    ws['C54'] = "=IF(C50>0,C50*(((1+(C53/100)/12)^(C51*12))-((1+(C53/100)/12)^C52))/(((1+(C53/100)/12)^(C51*12))-1),0)"
    ws['D54'] = "=IF(C54>0,C54-C50*(((1+(C53/100)/12)^(C51*12))-((1+(C53/100)/12)^(C52+12)))/(((1+(C53/100)/12)^(C51*12))-1),0)"
    ws['E54'] = "=MAX(0,C54-D54)"
    
    # LEASING
    ws['B56'] = "LEASING:"
    ws['C56'] = "Capital Pendiente"
    ws['D56'] = "Deuda CP"
    ws['E56'] = "Deuda LP"
    
    ws['B57'] = "Importe total:"
    ws['C57'] = "=Balance_Pasivo!B18"
    ws['B58'] = "Meses pendientes:"
    ws['C58'] = "=Balance_Pasivo!B19"
    ws['B59'] = "Cuota mensual:"
    ws['C59'] = "=Balance_Pasivo!B20"
    
    ws['B60'] = "Cálculo:"
    ws['C60'] = "=C57"
    ws['D60'] = "=IF(C58<=12,C57,C59*12)"
    ws['E60'] = "=MAX(0,C57-D60)"
    
    # Ajustar anchos de columna
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 20
    
    return ws

# Actualizar la referencia en crear_plantilla_v2
# Buscar la línea que llama a crear_balance_check y reemplazarla
