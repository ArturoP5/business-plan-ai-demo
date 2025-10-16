"""
Plantilla Excel DEFINITIVA - Versión completa y profesional
Compatible 100% con el sidebar de la aplicación
Con fórmulas, validaciones e instrucciones claras
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO

# Constantes de estilo
AZUL_OSCURO = "1E3A8A"
GRIS_CLARO = "F3F4F6"
VERDE = "10B981"

def crear_plantilla_definitiva():
    """
    Crea la plantilla Excel definitiva con todos los campos necesarios
    """
    wb = Workbook()
    año_actual = datetime.now().year
    
    # Estilos
    estilo_titulo = Font(bold=True, size=14, color="FFFFFF")
    fondo_titulo = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
    estilo_encabezado = Font(bold=True, size=11)
    fondo_encabezado = PatternFill(start_color=GRIS_CLARO, end_color=GRIS_CLARO, fill_type="solid")
    
    # HOJA 1: Instrucciones
    ws_instrucciones = wb.active
    ws_instrucciones.title = "INSTRUCCIONES"
    
    instrucciones = [
        ["PLANTILLA DE CAPTURA DE DATOS FINANCIEROS"],
        [""],
        ["📋 CÓMO USAR ESTA PLANTILLA:"],
        [""],
        ["1. LEA ESTAS INSTRUCCIONES COMPLETAS ANTES DE EMPEZAR"],
        ["2. Complete las hojas en el siguiente orden:"],
        ["   • Info_General: Datos básicos de la empresa"],
        ["   • PYL_Historico: Ventas y gastos de los últimos 3 años"],
        ["   • Balance_Activo: Activos de la empresa"],
        ["   • Balance_Pasivo: Pasivos y deudas"],
        ["   • Balance_Patrimonio: Capital y reservas"],
        ["   • Datos_Laborales: Información de empleados"],
        ["   • Lineas_Financiacion: Todas las líneas de crédito"],
        ["   • Parametros: Parámetros para proyecciones"],
        [""],
        ["⚠️ IMPORTANTE:"],
        ["• Use valores 0 donde no haya datos (no deje celdas vacías)"],
        ["• La depreciación debe ser NEGATIVA"],
        ["• El balance debe cuadrar: Activos = Pasivos + Patrimonio"],
        ["• Guarde el archivo como .xlsx"],
        [""],
        ["✅ VERIFICACIÓN:"],
        ["• En la hoja Balance_Check puede verificar si el balance cuadra"],
        ["• Los valores en amarillo se calculan automáticamente"],
        ["• Los valores en blanco debe introducirlos usted"]
    ]
    
    for row_idx, instruccion in enumerate(instrucciones, 1):
        if instruccion:
            ws_instrucciones.cell(row=row_idx, column=1, value=instruccion[0])
            if row_idx == 1:
                ws_instrucciones.cell(row=row_idx, column=1).font = Font(bold=True, size=16, color=AZUL_OSCURO)
            elif "📋" in instruccion[0] or "⚠️" in instruccion[0] or "✅" in instruccion[0]:
                ws_instrucciones.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
    
    ws_instrucciones.column_dimensions['A'].width = 80
    
    # Esta es la estructura base. Continuaremos agregando las demás hojas...
    
    return guardar_workbook_a_bytes(wb)

def guardar_workbook_a_bytes(wb):
    """Convierte el workbook a BytesIO para descarga"""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Función temporal para test
if __name__ == "__main__":
    print("✅ Módulo de plantilla definitiva creado")
    print("📋 Este archivo contendrá la plantilla completa y profesional")

def agregar_hoja_info_general(wb, año_actual):
    """Agrega la hoja de información general de la empresa"""
    ws = wb.create_sheet("Info_General")
    
    # Datos a capturar
    datos = [
        ["Campo", "Valor", "Instrucciones"],
        ["Nombre de la empresa", "", "Razón social completa"],
        ["Sector", "Tecnología", "Tecnología/Industrial/Retail/Servicios/Otro"],
        ["País", "España", "España/Francia/Portugal/Italia/Alemania/Reino Unido/Otro"],
        ["Año de Fundación", año_actual - 5, f"Entre 1900 y {año_actual}"],
        ["¿Empresa familiar?", "No", "Escriba: Sí o No"],
        ["¿Cuentas auditadas?", "Sí", "Escriba: Sí o No"],
        ["Moneda", "EUR", "EUR/USD/GBP"],
        ["", "", ""],
        ["--- DESCRIPCIÓN DEL NEGOCIO ---", "", "Sección nueva - Importante para análisis"],
        ["Descripción de la Actividad", "", "¿Qué hace la empresa? (máx 500 caracteres)"],
        ["Productos/Servicios Principales", "", "Lista sus principales productos o servicios"],
        ["Cuota de Mercado (%)", 0, "Porcentaje estimado en su segmento (0-100)"],
        ["Posicionamiento de Precios", "Medio", "Premium/Medio/Low-cost"],
        ["Ventajas Competitivas", "", "Principales diferenciadores vs competencia"],
        ["Clientes Objetivo", "", "Descripción del segmento de clientes"]
    ]
    
    # Escribir datos
    for row_idx, row_data in enumerate(datos, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Formato encabezados
            if row_idx == 1:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
            
            # Formato separadores
            if "---" in str(value):
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color=GRIS_CLARO, end_color=GRIS_CLARO, fill_type="solid")
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    
    return ws

def agregar_hoja_pyl_historico(wb, año_actual):
    """Agrega la hoja de P&L histórico con estructura correcta para importación"""
    ws = wb.create_sheet("Datos_Historicos_PYL")
    
    # Encabezados
    ws['A1'] = "Concepto"
    ws['B1'] = str(año_actual - 2)
    ws['C1'] = str(año_actual - 1)
    ws['D1'] = str(año_actual)
    
    # Conceptos del P&L
    conceptos = [
        "Ventas",
        "Costos Variables (%)",
        "Gastos de Personal",
        "Gastos Generales",
        "Gastos de Marketing"
    ]
    
    # Escribir conceptos
    for idx, concepto in enumerate(conceptos, 2):
        ws[f'A{idx}'] = concepto
        ws[f'B{idx}'] = 0
        ws[f'C{idx}'] = 0
        ws[f'D{idx}'] = 0
    
    # Formato especial para el porcentaje
    ws['B3'] = 40  # % por defecto
    ws['C3'] = "=B3"
    ws['D3'] = "=B3"
    
    # Agregar fila de EBITDA calculado
    ws['A8'] = "EBITDA (calculado)"
    ws['B8'] = "=B2*(1-B3/100)-B4-B5-B6"
    ws['C8'] = "=C2*(1-C3/100)-C4-C5-C6"
    ws['D8'] = "=D2*(1-D3/100)-D4-D5-D6"
    
    # Agregar margen EBITDA
    ws['A9'] = "Margen EBITDA (%)"
    ws['B9'] = "=IF(B2>0,B8/B2*100,0)"
    ws['C9'] = "=IF(C2>0,C8/C2*100,0)"
    ws['D9'] = "=IF(D2>0,D8/D2*100,0)"
    
    # Formato
    for cell in [ws['A1'], ws['B1'], ws['C1'], ws['D1']]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
    
    # Resaltar celdas calculadas
    for row in [8, 9]:
        for col in ['B', 'C', 'D']:
            ws[f'{col}{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            ws[f'{col}{row}'].font = Font(italic=True)
    
    # Anchos de columna
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    return ws

def agregar_hojas_balance(wb, año_actual):
    """Agrega las 3 hojas del balance con fórmulas de cuadre"""
    
    # HOJA BALANCE - ACTIVO
    ws_activo = wb.create_sheet("Balance_Activo")
    
    activos = [
        ["Concepto", f"Valor {año_actual}", "Notas"],
        ["--- ACTIVO CORRIENTE ---", "", ""],
        ["Tesorería", 0, "Efectivo y equivalentes"],
        ["Inversiones financieras CP", 0, "Inversiones < 1 año"],
        ["Clientes", 0, "Cuentas por cobrar"],
        ["Inventario", 0, "Stock de productos"],
        ["Otros deudores", 0, "Otros deudores comerciales"],
        ["Administración Pública deudora", 0, "Hacienda deudora, devoluciones"],
        ["Gastos anticipados", 0, "Seguros, alquileres prepagados"],
        ["Activos por impuesto diferido CP", 0, "Créditos fiscales corto plazo"],
        ["TOTAL ACTIVO CORRIENTE", "=SUM(B3:B10)", "Calculado automáticamente"],
        ["", "", ""],
        ["--- ACTIVO NO CORRIENTE ---", "", ""],
        ["Activo fijo bruto", 0, "Inmuebles, maquinaria, equipos"],
        ["Depreciación acumulada", 0, "DEBE SER NEGATIVO"],
        ["Activo fijo neto", "=B14+B15", "Calculado automáticamente"],
        ["Activos intangibles brutos", 0, "Software, marcas, patentes"],
        ["Amortización acumulada intangibles", 0, "DEBE SER NEGATIVO"],
        ["Activos intangibles netos", "=B17+B18", "Calculado automáticamente"],
        ["Inversiones financieras LP", 0, "Inversiones > 1 año"],
        ["Créditos a largo plazo", 0, "Préstamos concedidos LP"],
        ["Fianzas y depósitos", 0, "Fianzas constituidas"],
        ["Activos por impuesto diferido LP", 0, "Créditos fiscales largo plazo"],
        ["TOTAL ACTIVO NO CORRIENTE", "=B16+B19+SUM(B20:B23)", "Calculado automáticamente"],
        ["", "", ""],
        ["TOTAL ACTIVOS", "=B11+B24", "ESTE VALOR DEBE COINCIDIR CON TOTAL PASIVO + PATRIMONIO"]
    ]
    
    escribir_datos_hoja(ws_activo, activos)
    formatear_totales(ws_activo, [11, 24, 26])
    resaltar_celdas_formula(ws_activo, [11, 16, 19, 24, 26])
    
    # HOJA BALANCE - PASIVO
    ws_pasivo = wb.create_sheet("Balance_Pasivo")
    
    pasivos = [
        ["Concepto", f"Valor {año_actual}", "Notas"],
        ["--- PASIVO CORRIENTE ---", "", ""],
        ["Proveedores", 0, "Facturas pendientes de pago"],
        ["Acreedores por servicios", 0, "Otros acreedores comerciales"],
        ["Anticipos de clientes", 0, "Cobros anticipados"],
        ["Remuneraciones pendientes", 0, "Salarios pendientes"],
        ["Administración Pública acreedora", 0, "IRPF, SS pendientes"],
        ["Provisiones CP", 0, "Provisiones varias"],
        ["Deuda financiera CP", 0, "Préstamos y líneas < 1 año"],
        ["TOTAL PASIVO CORRIENTE", "=SUM(B3:B9)", "Calculado automáticamente"],
        ["", "", ""],
        ["--- PASIVO NO CORRIENTE ---", "", ""],
        ["Préstamos LP - Importe original", 0, "Importe inicial del préstamo"],
        ["Préstamos LP - Años transcurridos", 0, "Años ya pagados"],
        ["Hipotecas - Importe original", 0, "Importe inicial hipoteca"],
        ["Hipotecas - Meses transcurridos", 0, "Meses ya pagados"],
        ["Leasing - Importe total", 0, "Valor total del bien"],
        ["Leasing - Meses pendientes", 0, "Meses que quedan"],
        ["Otros préstamos LP", 0, "Otros préstamos bancarios"],
        ["Provisiones para riesgos LP", 0, "Contingencias LP"],
        ["Pasivos por impuesto diferido", 0, "Pasivos fiscales diferidos"],
        ["TOTAL PASIVO NO CORRIENTE", "=SUM(B13,B15,B17:B21)", "Calculado automáticamente"],
        ["", "", ""],
        ["TOTAL PASIVOS", "=B10+B22", "Suma pasivo corriente + no corriente"]
    ]
    
    escribir_datos_hoja(ws_pasivo, pasivos)
    formatear_totales(ws_pasivo, [10, 22, 24])
    resaltar_celdas_formula(ws_pasivo, [10, 22, 24])
    
    # HOJA BALANCE - PATRIMONIO
    ws_patrimonio = wb.create_sheet("Balance_Patrimonio")
    
    patrimonio = [
        ["Concepto", f"Valor {año_actual}", "Notas"],
        ["Capital social", 0, "Capital desembolsado"],
        ["Prima de emisión", 0, "Prima sobre nominal"],
        ["Reserva legal", 0, "10% del capital social"],
        ["Otras reservas", 0, "Reservas voluntarias"],
        ["Resultados acumulados", 0, "Beneficios años anteriores"],
        ["Resultado del ejercicio", 0, "Resultado año actual"],
        ["Ajustes por cambios de valor", 0, "Ajustes valoración"],
        ["", "", ""],
        ["TOTAL PATRIMONIO NETO", "=SUM(B2:B8)", "ACTIVOS - PASIVOS = PATRIMONIO"]
    ]
    
    escribir_datos_hoja(ws_patrimonio, patrimonio)
    formatear_totales(ws_patrimonio, [10])
    resaltar_celdas_formula(ws_patrimonio, [10])
    
    return ws_activo, ws_pasivo, ws_patrimonio

def escribir_datos_hoja(ws, datos):
    """Escribe los datos en la hoja"""
    for row_idx, row_data in enumerate(datos, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Formato encabezados
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
            
            # Formato separadores
            elif "---" in str(value):
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=GRIS_CLARO, end_color=GRIS_CLARO, fill_type="solid")
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40

def formatear_totales(ws, filas):
    """Aplica formato especial a las filas de totales"""
    for fila in filas:
        ws[f'A{fila}'].font = Font(bold=True, size=11)
        ws[f'B{fila}'].font = Font(bold=True, size=11)

def resaltar_celdas_formula(ws, filas):
    """Resalta las celdas con fórmulas"""
    for fila in filas:
        ws[f'B{fila}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

def agregar_hoja_datos_laborales(wb):
    """Agrega la hoja de datos laborales"""
    ws = wb.create_sheet("Datos_Laborales")
    
    datos = [
        ["Concepto", "Valor", "Notas"],
        ["--- PLANTILLA ACTUAL ---", "", ""],
        ["Número de empleados", 0, "Total empleados actuales"],
        ["Coste medio por empleado (€/año)", 0, "Salario bruto + cargas sociales"],
        ["Antigüedad media (años)", 0, "Media de años en la empresa"],
        ["Rotación anual (%)", 0, "Porcentaje de rotación esperada"],
        ["", "", ""],
        ["--- REESTRUCTURACIÓN (si aplica) ---", "", ""],
        ["% plantilla afectada", 0, "Dejar 0 si no hay reestructuración"],
        ["Días indemnización por año", 0, "20, 33, 45 o especificar"],
        ["", "", ""],
        ["--- PROVISIONES M&A ---", "", ""],
        ["Provisión para litigios", 0, "Contingencias laborales"],
        ["Provisión fiscal", 0, "Contingencias fiscales"]
    ]
    
    escribir_datos_hoja(ws, datos)
    return ws

def agregar_hoja_lineas_financiacion(wb):
    """Agrega la hoja de líneas de financiación con múltiples filas por tipo"""
    ws = wb.create_sheet("Lineas_Financiacion")
    
    # Encabezados
    encabezados = ["Tipo de Línea", "Banco", "Límite", "Dispuesto", "Tipo (%)", "Comisión (%)"]
    for col_idx, encabezado in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col_idx, value=encabezado)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
    
    # Tipos de líneas (3 filas para cada tipo)
    tipos_lineas = [
        "Póliza de Crédito",
        "Póliza Crédito Stock",
        "Descuento Comercial",
        "Anticipo de Facturas",
        "Factoring con Recurso",
        "Factoring sin Recurso",
        "Confirming Proveedores",
        "Pagarés Empresa",
        "Crédito Importación"
    ]
    
    row_idx = 2
    for tipo in tipos_lineas:
        # Primera línea con el nombre del tipo
        ws.cell(row=row_idx, column=1, value=tipo)
        for col in range(2, 7):
            ws.cell(row=row_idx, column=col, value=0 if col > 2 else "")
        row_idx += 1
        
        # Dos líneas adicionales vacías para el mismo tipo
        for _ in range(2):
            ws.cell(row=row_idx, column=1, value="")
            for col in range(2, 7):
                ws.cell(row=row_idx, column=col, value=0 if col > 2 else "")
            row_idx += 1
    
    # Líneas adicionales para otros tipos
    ws.cell(row=row_idx, column=1, value="--- OTRAS LÍNEAS ---")
    for col in range(1, 7):
        ws.cell(row=row_idx, column=col).fill = PatternFill(start_color=GRIS_CLARO, end_color=GRIS_CLARO, fill_type="solid")
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    
    return ws

def agregar_hoja_parametros(wb):
    """Agrega la hoja de parámetros de proyección"""
    ws = wb.create_sheet("Proyecciones_Parametros")
    
    datos = [
        ["Parámetro", "Valor", "Referencia"],
        ["--- PARÁMETROS OPERATIVOS ---", "", ""],
        ["Días de cobro (DSO)", 60, "Media sector: 60-90 días"],
        ["Días de pago (DPO)", 45, "Media sector: 30-60 días"],
        ["Días de inventario", 30, "Según tipo de negocio"],
        ["", "", ""],
        ["--- CRECIMIENTO DE VENTAS ---", "", ""],
        ["Crecimiento Año 1 (%)", 10, "Sobre año actual"],
        ["Crecimiento Año 2 (%)", 8, "Sobre año 1"],
        ["Crecimiento Año 3 (%)", 6, "Sobre año 2"],
        ["Crecimiento Año 4 (%)", 5, "Sobre año 3"],
        ["Crecimiento Año 5 (%)", 4, "Sobre año 4"],
        ["", "", ""],
        ["--- INVERSIONES (CAPEX) ---", "", ""],
        ["CAPEX Año 1", 0, "Inversión en activos fijos"],
        ["CAPEX Año 2", 0, "Inversión en activos fijos"],
        ["CAPEX Año 3", 0, "Inversión en activos fijos"],
        ["CAPEX Año 4", 0, "Inversión en activos fijos"],
        ["CAPEX Año 5", 0, "Inversión en activos fijos"],
        ["", "", ""],
        ["--- EVOLUCIÓN PLANTILLA ---", "", ""],
        ["Nuevos empleados Año 1", 0, "Incremento neto de plantilla"],
        ["Nuevos empleados Año 2", 0, "Incremento neto de plantilla"],
        ["Nuevos empleados Año 3", 0, "Incremento neto de plantilla"],
        ["Nuevos empleados Año 4", 0, "Incremento neto de plantilla"],
        ["Nuevos empleados Año 5", 0, "Incremento neto de plantilla"]
    ]
    
    escribir_datos_hoja(ws, datos)
    return ws

def agregar_hoja_verificacion(wb):
    """Agrega una hoja de verificación del balance"""
    ws = wb.create_sheet("Balance_Check")
    
    ws['A1'] = "VERIFICACIÓN DE BALANCE"
    ws['A1'].font = Font(bold=True, size=14, color=AZUL_OSCURO)
    
    ws['A3'] = "RESUMEN"
    ws['A3'].font = Font(bold=True, size=12)
    
    ws['A5'] = "Total Activos:"
    ws['B5'] = "='Balance_Activo'!B26"
    
    ws['A6'] = "Total Pasivos:"
    ws['B6'] = "='Balance_Pasivo'!B24"
    
    ws['A7'] = "Total Patrimonio:"
    ws['B7'] = "='Balance_Patrimonio'!B10"
    
    ws['A9'] = "Pasivos + Patrimonio:"
    ws['B9'] = "=B6+B7"
    ws['B9'].font = Font(bold=True)
    
    ws['A11'] = "DIFERENCIA (debe ser 0):"
    ws['B11'] = "=B5-B9"
    ws['B11'].font = Font(bold=True, size=12)
    
    # Formato condicional manual
    ws['A13'] = "ESTADO:"
    ws['B13'] = '=IF(ABS(B11)<1,"✅ BALANCE CUADRADO","❌ BALANCE DESCUADRADO")'
    ws['B13'].font = Font(bold=True, size=12)
    
    # Instrucciones
    ws['A15'] = "INSTRUCCIONES SI EL BALANCE NO CUADRA:"
    ws['A15'].font = Font(bold=True)
    
    instrucciones = [
        "1. Revise que todos los activos estén registrados",
        "2. Verifique que la depreciación sea negativa",
        "3. Compruebe que todos los pasivos estén incluidos",
        "4. Asegúrese de que el patrimonio esté completo",
        "5. Revise los cálculos manuales"
    ]
    
    for idx, instruccion in enumerate(instrucciones, 17):
        ws[f'A{idx}'] = instruccion
    
    # Formato
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    
    # Resaltar celdas importantes
    ws['B11'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    ws['B13'].fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    
    return ws

# Actualizar la función principal para incluir todas las hojas
def crear_plantilla_definitiva():
    """
    Crea la plantilla Excel definitiva con todos los campos necesarios
    """
    wb = Workbook()
    año_actual = datetime.now().year
    
    # Eliminar la hoja por defecto
    wb.remove(wb.active)
    
    # Agregar todas las hojas en orden
    agregar_hoja_instrucciones(wb)
    agregar_hoja_info_general(wb, año_actual)
    agregar_hoja_pyl_historico(wb, año_actual)
    agregar_hojas_balance(wb, año_actual)
    agregar_hoja_datos_laborales(wb)
    agregar_hoja_lineas_financiacion(wb)
    agregar_hoja_parametros(wb)
    agregar_hoja_verificacion(wb)
    
    return guardar_workbook_a_bytes(wb)

def agregar_hoja_instrucciones(wb):
    """Agrega la hoja de instrucciones iniciales"""
    ws = wb.create_sheet("INSTRUCCIONES")
    
    instrucciones = [
        ["PLANTILLA DE CAPTURA DE DATOS FINANCIEROS - BUSINESS PLAN IA"],
        [""],
        ["📋 CÓMO USAR ESTA PLANTILLA:"],
        [""],
        ["1. LEA ESTAS INSTRUCCIONES COMPLETAS"],
        ["2. Complete las hojas en orden:"],
        ["   • Info_General → Información básica de la empresa"],
        ["   • Datos_Historicos_PYL → Ventas y gastos últimos 3 años"],
        ["   • Balance_Activo → Todos los activos"],
        ["   • Balance_Pasivo → Todas las deudas"],
        ["   • Balance_Patrimonio → Capital y reservas"],
        ["   • Datos_Laborales → Información de empleados"],
        ["   • Lineas_Financiacion → Líneas de crédito activas"],
        ["   • Proyecciones_Parametros → Parámetros para proyecciones"],
        [""],
        ["⚠️ MUY IMPORTANTE:"],
        ["• Use 0 donde no haya datos (NO deje celdas vacías)"],
        ["• La depreciación/amortización debe ser NEGATIVA"],
        ["• Las celdas amarillas se calculan automáticamente"],
        ["• Las celdas blancas requieren su entrada"],
        ["• El balance debe cuadrar (verifique en Balance_Check)"],
        [""],
        ["✅ ANTES DE IMPORTAR:"],
        ["• Revise la hoja Balance_Check"],
        ["• Asegúrese de que muestre 'BALANCE CUADRADO'"],
        ["• Guarde el archivo como .xlsx"],
        ["• NO modifique los nombres de las hojas"],
        [""],
        ["📧 SOPORTE:"],
        ["• Los valores en las columnas 'Notas' son informativos"],
        ["• Complete SOLO las columnas 'Valor'"],
        ["• Si tiene dudas, use los valores por defecto"]
    ]
    
    for row_idx, instruccion in enumerate(instrucciones, 1):
        if instruccion:
            cell = ws.cell(row=row_idx, column=1, value=instruccion[0])
            if row_idx == 1:
                cell.font = Font(bold=True, size=16, color=AZUL_OSCURO)
            elif any(x in instruccion[0] for x in ["📋", "⚠️", "✅", "📧"]):
                cell.font = Font(bold=True, size=12)
    
    ws.column_dimensions['A'].width = 80
    return ws

# Test final
if __name__ == "__main__":
    excel = crear_plantilla_definitiva()
    with open("plantilla_definitiva_test.xlsx", "wb") as f:
        f.write(excel.getvalue())
    print("✅ Plantilla definitiva completa creada")
    print("📊 Incluye:")
    print("  • Instrucciones claras")
    print("  • Todos los campos del sidebar")
    print("  • Fórmulas automáticas")
    print("  • Verificación de balance")
    print("  • Valores por defecto = 0")
