"""
Plantilla Excel COMPLETA compatible con leer_excel_datos
Incluye las nuevas líneas de financiación y descripción del negocio
"""

import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment

# Colores corporativos
COLOR_PRIMARIO = "1E3A8A"
COLOR_SECUNDARIO = "3B82F6"  
COLOR_FONDO = "F3F4F6"

def crear_plantilla_completa_compatible():
    """
    Crea plantilla Excel 100% compatible con leer_excel_datos
    Incluye TODAS las hojas requeridas con los nombres exactos
    """
    output = BytesIO()
    año_actual = datetime.now().year
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Las hojas se agregarán en las siguientes partes
        pass
    
    output.seek(0)
    return output

def crear_hoja_info_general(writer, año_actual):
    """Hoja 1: Información General con nuevos campos de descripción del negocio"""
    df = pd.DataFrame({
        'Campo': [
            'Nombre de la empresa',
            'Sector',
            'País',
            'Año de Fundación',
            '¿Empresa familiar?',
            '¿Cuentas auditadas?',
            'Moneda',
            '',
            '--- NUEVA SECCIÓN: DESCRIPCIÓN DEL NEGOCIO ---',
            'Descripción de la Actividad',
            'Productos/Servicios Principales',
            'Cuota de Mercado (%)',
            'Posicionamiento de Precios',
            'Ventajas Competitivas',
            'Clientes Objetivo'
        ],
        'Valor': [''] * 15,
        'Instrucciones': [
            'Razón social completa',
            'Tecnología/E-commerce/Retail/Servicios/Industrial/Otro',
            'España/Francia/Portugal/Italia/Alemania/Reino Unido/Otro',
            f'Entre 1900 y {año_actual}',
            'Sí/No',
            'Sí/No', 
            'EUR/USD/GBP',
            '',
            'NUEVOS CAMPOS para análisis con IA:',
            '¿Qué hace la empresa?',
            'Lista de productos/servicios principales',
            'Porcentaje estimado en su segmento',
            'Premium/Medio/Low-cost',
            'Principales diferenciadores',
            'Segmento de clientes objetivo'
        ]
    })
    df.to_excel(writer, sheet_name='Informacion General', index=False)

def crear_hoja_pyl_historico(writer, año_actual):
    """Hoja 2: Datos Históricos PYL - ESTRUCTURA CORRECTA"""
    df = pd.DataFrame({
        "Concepto": [
            "Ventas",
            "Costos Variables (%)",
            "Gastos de Personal",
            "Gastos Generales",
            "Gastos de Marketing"
        ],
        str(año_actual-2): ["", "", "", "", ""],
        str(año_actual-1): ["", "", "", "", ""],
        str(año_actual): ["", "", "", "", ""]
    })
    df.to_excel(writer, sheet_name="Datos Históricos PYL", index=False)
def crear_hoja_balance_activo(writer, año_actual):
    """Hoja 3: Balance - Activo"""
    df = pd.DataFrame({
        'Concepto': [
            '--- ACTIVO CORRIENTE ---',
            'Tesorería',
            'Inversiones financieras CP',
            'Clientes',
            'Inventario',
            'Otros deudores',
            'Administración Pública deudora',
            'Gastos anticipados',
            'Activos por impuesto diferido CP',
            '',
            '--- ACTIVO NO CORRIENTE ---',
            'Activo fijo bruto',
            'Depreciación acumulada',
            'Activos intangibles brutos',
            'Amortización acumulada intangibles',
            'Inversiones financieras LP',
            'Créditos a largo plazo',
            'Fianzas y depósitos',
            'Activos por impuesto diferido LP'
        ],
        f'Valor {año_actual}': [''] * 19,
        'Notas': [
            '',
            'Efectivo y bancos',
            'Inversiones < 1 año',
            'Cuentas por cobrar',
            'Stock de productos',
            'Otros deudores',
            'Hacienda deudora',
            'Pagos anticipados',
            'Créditos fiscales CP',
            '',
            '',
            'Inmuebles, maquinaria, equipos',
            'Usar signo NEGATIVO',
            'Software, marcas, patentes',
            'Usar signo NEGATIVO',
            'Inversiones > 1 año',
            'Préstamos concedidos LP',
            'Fianzas constituidas',
            'Créditos fiscales LP'
        ]
    })
    df.to_excel(writer, sheet_name='Balance - Activo', index=False)

def crear_hoja_balance_pasivo(writer, año_actual):
    """Hoja 4: Balance - Pasivo"""
    df = pd.DataFrame({
        'Concepto': [
            '--- PASIVO CORRIENTE ---',
            'Proveedores',
            'Acreedores por servicios',
            'Anticipos de clientes',
            'Remuneraciones pendientes',
            'Administración Pública acreedora',
            'Provisiones CP',
            'Deuda financiera CP (ver hoja Líneas)',
            '',
            '--- PASIVO NO CORRIENTE ---',
            'Préstamos LP - Importe original',
            'Préstamos LP - Años transcurridos',
            'Hipotecas - Importe original',
            'Hipotecas - Meses transcurridos',
            'Leasing - Importe total',
            'Leasing - Meses pendientes',
            'Otros préstamos LP',
            'Provisiones para riesgos LP',
            'Pasivos por impuesto diferido'
        ],
        f'Valor {año_actual}': [''] * 19,
        'Notas': [
            '',
            'Facturas pendientes de pago',
            'Otros acreedores',
            'Cobros anticipados',
            'Nóminas pendientes',
            'IRPF, SS pendientes',
            'Provisiones varias CP',
            'Suma líneas circulante',
            '',
            '',
            'Importe inicial del préstamo',
            'Años ya pagados',
            'Importe inicial hipoteca',
            'Meses ya pagados',
            'Valor total del bien',
            'Meses que quedan por pagar',
            'Otros préstamos bancarios',
            'Contingencias LP',
            'Pasivos fiscales diferidos'
        ]
    })
    df.to_excel(writer, sheet_name='Balance - Pasivo', index=False)

def crear_hoja_balance_patrimonio(writer, año_actual):
    """Hoja 5: Balance - Patrimonio"""
    df = pd.DataFrame({
        'Concepto': [
            'Capital social',
            'Prima de emisión',
            'Reserva legal',
            'Otras reservas',
            'Resultados acumulados',
            'Resultado del ejercicio',
            'Ajustes por cambios de valor'
        ],
        f'Valor {año_actual}': [''] * 7,
        'Notas': [
            'Capital desembolsado',
            'Prima sobre nominal',
            'Reserva legal (10% capital)',
            'Reservas voluntarias',
            'Resultados años anteriores',
            'Resultado año actual',
            'Ajustes valoración'
        ]
    })
    df.to_excel(writer, sheet_name='Balance - Patrimonio', index=False)

def crear_hoja_datos_laborales(writer):
    """Hoja 6: Datos Laborales"""
    df = pd.DataFrame({
        'Concepto': [
            'Número de empleados',
            'Coste medio por empleado (€/año)',
            'Antigüedad media (años)',
            'Rotación anual (%)',
            '',
            '--- REESTRUCTURACIÓN ---',
            '% plantilla afectada',
            'Días indemnización por año',
            '',
            '--- PROVISIONES M&A ---',
            'Provisión para litigios',
            'Provisión fiscal'
        ],
        'Valor': [''] * 12,
        'Notas': [
            'Total empleados',
            'Salario + SS',
            'Media años',
            'Tasa rotación',
            '',
            '',
            'Si aplica',
            '20, 33, 45',
            '',
            '',
            'Litigios laborales',
            'Contingencias'
        ]
    })
    df.to_excel(writer, sheet_name='Datos Laborales', index=False)

def crear_hoja_lineas_financiacion(writer):
    """Hoja 7: Líneas Financiación - Permite múltiples líneas del mismo tipo"""
    
    tipos_lineas = [
        "Póliza de Crédito",
        "Póliza Crédito Stock",
        "Descuento Comercial",
        "Anticipo de Facturas",
        "Factoring con Recurso",
        "Factoring sin Recurso",
        "Confirming Proveedores",
        "Pagarés Empresa",
        "Crédito Importación"
    ]
    
    data = []
    data.append(["--- Puede agregar múltiples líneas del mismo tipo (diferentes bancos) ---", "", "", "", "", ""])
    
    for tipo in tipos_lineas:
        data.append([tipo, "", 0, 0, 0, 0])
        for i in range(2):  # 2 líneas adicionales vacías para cada tipo
            data.append(["", "", 0, 0, 0, 0])
    
    data.append(["", "", "", "", "", ""])
    data.append(["--- OTRAS LÍNEAS (especificar tipo) ---", "", "", "", "", ""])
    for i in range(5):
        data.append(["", "", 0, 0, 0, 0])
    
    df = pd.DataFrame(data, columns=["Tipo de Línea", "Banco/Entidad", "Límite", "Dispuesto", "Tipo (%)", "Comisión (%)"])
    df.to_excel(writer, sheet_name="Líneas Financiación", index=False)
def crear_hoja_proyecciones(writer):
    """Hoja 8: Proyecciones y Parámetros"""
    df = pd.DataFrame({
        'Parámetro': [
            '--- PARÁMETROS OPERATIVOS ---',
            'Días de cobro (DSO)',
            'Días de pago (DPO)',
            'Días de inventario',
            '',
            '--- CRECIMIENTO VENTAS ---',
            'Crecimiento Año 1 (%)',
            'Crecimiento Año 2 (%)',
            'Crecimiento Año 3 (%)',
            'Crecimiento Año 4 (%)',
            'Crecimiento Año 5 (%)',
            '',
            '--- CAPEX ---',
            'CAPEX Año 1',
            'CAPEX Año 2',
            'CAPEX Año 3',
            'CAPEX Año 4',
            'CAPEX Año 5',
            '',
            '--- PLANTILLA ---',
            'Nuevos empleados Año 1',
            'Nuevos empleados Año 2',
            'Nuevos empleados Año 3',
            'Nuevos empleados Año 4',
            'Nuevos empleados Año 5'
        ],
        'Valor': [''] * 25,
        'Referencia': [
            '',
            'Media: 60-90',
            'Media: 30-60',
            'Según sector',
            '',
            '',
            'Sobre año anterior',
            'Sobre año 1',
            'Sobre año 2',
            'Sobre año 3',
            'Sobre año 4',
            '',
            '',
            'Inversión €',
            'Inversión €',
            'Inversión €',
            'Inversión €',
            'Inversión €',
            '',
            '',
            'Incremento neto',
            'Incremento neto',
            'Incremento neto',
            'Incremento neto',
            'Incremento neto'
        ]
    })
    df.to_excel(writer, sheet_name='Proyecciones y Parámetros', index=False)

# Actualizar la función principal
def crear_plantilla_completa():
    """
    Crea plantilla Excel completa y compatible con leer_excel_datos
    """
    output = BytesIO()
    año_actual = datetime.now().year
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Crear todas las hojas
        crear_hoja_info_general(writer, año_actual)
        crear_hoja_pyl_historico(writer, año_actual)
        crear_hoja_balance_activo(writer, año_actual)
        crear_hoja_balance_pasivo(writer, año_actual)
        crear_hoja_balance_patrimonio(writer, año_actual)
        crear_hoja_datos_laborales(writer)
        crear_hoja_lineas_financiacion(writer)
        crear_hoja_proyecciones(writer)
        
        # Aplicar formato profesional
        aplicar_formato_profesional(writer.book)
    
    output.seek(0)
    return output

def aplicar_formato_profesional(workbook):
    """Aplica formato McKinsey a todas las hojas"""
    
    header_fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    separator_fill = PatternFill(start_color=COLOR_FONDO, end_color=COLOR_FONDO, fill_type='solid')
    separator_font = Font(bold=True, size=10, color=COLOR_PRIMARIO)
    
    for sheet in workbook.worksheets:
        # Configurar anchos
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 25
        if sheet.max_column >= 3:
            sheet.column_dimensions['C'].width = 35
        
        # Formato encabezados (primera fila)
        for cell in sheet[1]:
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Identificar y formatear separadores
        for row in sheet.iter_rows(min_row=2):
            if row[0].value and '---' in str(row[0].value):
                for cell in row:
                    cell.fill = separator_fill
                    cell.font = separator_font
        
        # Congelar primera fila
        sheet.freeze_panes = 'A2'
