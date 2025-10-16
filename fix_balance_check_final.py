from openpyxl import load_workbook

wb = load_workbook('excel_referencias_corregidas.xlsx')
ws = wb['Balance_Check']

# Limpiar la zona de cálculos (filas 42-60)
for row in range(42, 61):
    for col in range(2, 6):  # Columnas B a E
        ws.cell(row=row, column=col).value = None

# SECCIÓN PRÉSTAMO (filas 42-47)
ws['B42'] = "PRÉSTAMO BANCARIO:"
ws['C42'] = "Capital Pendiente"
ws['D42'] = "Deuda CP"
ws['E42'] = "Deuda LP"

ws['B43'] = "Importe original:"
ws['C43'] = "=Balance_Pasivo!B9"
ws['B44'] = "Plazo (años):"
ws['C44'] = "=Balance_Pasivo!B11"
ws['B45'] = "Años transcurridos:"
ws['C45'] = "=Balance_Pasivo!B10"
ws['B46'] = "Tipo interés (%):"
ws['C46'] = "=Balance_Pasivo!B12"

# Fórmula correcta para capital pendiente (amortización francesa)
ws['B47'] = "Cálculo:"
ws['C47'] = "=IF(C43>0,C43*(((1+(C46/100)/12)^(C44*12))-((1+(C46/100)/12)^(C45*12)))/(((1+(C46/100)/12)^(C44*12))-1),0)"
# CP: capital de los próximos 12 meses
ws['D47'] = "=IF(C47>0,C47-C43*(((1+(C46/100)/12)^(C44*12))-((1+(C46/100)/12)^((C45+1)*12)))/(((1+(C46/100)/12)^(C44*12))-1),0)"
# LP: resto
ws['E47'] = "=MAX(0,C47-D47)"

# SECCIÓN HIPOTECA (filas 49-54)
ws['B49'] = "HIPOTECA:"
ws['C49'] = "Capital Pendiente"
ws['D49'] = "Deuda CP"
ws['E49'] = "Deuda LP"

ws['B50'] = "Importe original:"
ws['C50'] = "=Balance_Pasivo!B14"
ws['B51'] = "Plazo (años):"
ws['C51'] = "=Balance_Pasivo!B16"
ws['B52'] = "Meses transcurridos:"
ws['C52'] = "=Balance_Pasivo!B15"
ws['B53'] = "Tipo interés (%):"
ws['C53'] = "=Balance_Pasivo!B17"

# Fórmula hipoteca
ws['B54'] = "Cálculo:"
ws['C54'] = "=IF(C50>0,C50*(((1+(C53/100)/12)^(C51*12))-((1+(C53/100)/12)^C52))/(((1+(C53/100)/12)^(C51*12))-1),0)"
# CP hipoteca
ws['D54'] = "=IF(C54>0,C54-C50*(((1+(C53/100)/12)^(C51*12))-((1+(C53/100)/12)^(C52+12)))/(((1+(C53/100)/12)^(C51*12))-1),0)"
# LP hipoteca
ws['E54'] = "=MAX(0,C54-D54)"

# SECCIÓN LEASING (filas 56-59)
ws['B56'] = "LEASING:"
ws['C56'] = "Capital Pendiente"
ws['D56'] = "Deuda CP"
ws['E56'] = "Deuda LP"

ws['B57'] = "Importe total:"
ws['C57'] = "=Balance_Pasivo!B18"
ws['B58'] = "Meses pendientes:"
ws['C58'] = "=Balance_Pasivo!B19"
ws['B59'] = "Cuota mensual:"
ws['C59'] = "=Balance_Pasivo!B20"

# En leasing, el pendiente es el total
ws['C59'] = "=C57"  # Capital pendiente = total
ws['D59'] = "=IF(C58<=12,C57,MIN(C57,C59*12))"  # CP
ws['E59'] = "=MAX(0,C57-D59)"  # LP

# ACTUALIZAR LAS REFERENCIAS EN EL BALANCE
# Préstamo
ws['G14'] = "=D47"  # CP
ws['G20'] = "=E47"  # LP

# Hipoteca  
ws['G15'] = "=D54"  # CP
ws['G21'] = "=E54"  # LP

# Leasing
ws['G16'] = "=D59"  # CP
ws['G22'] = "=E59"  # LP

wb.save('excel_balance_final.xlsx')
print("✅ Balance_Check arreglado definitivamente")
print("Las fórmulas ahora apuntan a las celdas correctas")
